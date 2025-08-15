import pandas as pd
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload
from app.models.company import Company
from app.models.aum_snapshot import AUMSnapshot
from app.models.scrape_log import ScrapeLog
from app.utils.unit_converter import format_currency
from datetime import datetime
from typing import List, Dict
import logging
import io

logger = logging.getLogger(__name__)

class ExcelExporter:
    def __init__(self):
        self.output_filename = f"aum_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    async def export_results(self, db: AsyncSession) -> bytes:
        """Exporta todos os resultados para Excel"""
        
        # Buscar dados das empresas com AUM
        companies_data = await self._get_companies_data(db)
        
        # Criar DataFrames
        main_df = await self._create_main_dataframe(companies_data)
        logs_df = await self._create_logs_dataframe(db)
        summary_df = await self._create_summary_dataframe(companies_data)
        
        # Criar arquivo Excel em memória
        excel_buffer = io.BytesIO()
        
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            # Planilha principal
            main_df.to_excel(writer, sheet_name='Resultados AUM', index=False)
            
            # Planilha de logs
            logs_df.to_excel(writer, sheet_name='Logs de Scraping', index=False)
            
            # Planilha de resumo
            summary_df.to_excel(writer, sheet_name='Resumo', index=False)
            
            # Formatar planilhas
            await self._format_excel_sheets(writer)
        
        excel_buffer.seek(0)
        return excel_buffer.getvalue()
    
    async def _get_companies_data(self, db: AsyncSession) -> List[Dict]:
        """Busca dados das empresas com seus AUMs"""
        
        result = await db.execute(
            select(Company).options(
                selectinload(Company.aum_snapshots),
                selectinload(Company.scrape_logs)
            )
        )
        
        companies = result.scalars().all()
        companies_data = []
        
        for company in companies:
            # Pegar o AUM mais recente
            latest_aum = None
            if company.aum_snapshots:
                latest_aum = max(company.aum_snapshots, key=lambda x: x.extracted_at)
            
            companies_data.append({
                "company": company,
                "latest_aum": latest_aum,
                "scrape_logs": company.scrape_logs
            })
        
        return companies_data
    
    async def _create_main_dataframe(self, companies_data: List[Dict]) -> pd.DataFrame:
        """Cria DataFrame principal com resultados"""
        
        data = []
        
        for item in companies_data:
            company = item["company"]
            latest_aum = item["latest_aum"]
            scrape_logs = item["scrape_logs"]
            
            # Contar status dos scrapes
            successful_scrapes = len([log for log in scrape_logs if log.status == "success"])
            failed_scrapes = len([log for log in scrape_logs if log.status == "failed"])
            
            # URLs que foram scrapadas
            scraped_urls = [log.url for log in scrape_logs if log.status == "success"]
            
            row = {
                "Empresa": company.name,
                "Site": company.url_site or "N/A",
                "LinkedIn": company.url_linkedin or "N/A",
                "Instagram": company.url_instagram or "N/A",
                "Twitter/X": company.url_x or "N/A",
                "AUM Encontrado": latest_aum.aum_raw_text if latest_aum else "NAO_DISPONIVEL",
                "AUM Normalizado": latest_aum.aum_normalized if latest_aum else None,
                "AUM Formatado": format_currency(latest_aum.aum_normalized) if latest_aum and latest_aum.aum_normalized else "N/A",
                "Fonte": latest_aum.source_url if latest_aum else "N/A",
                "Método Extração": latest_aum.extraction_method if latest_aum else "N/A",
                "Confiança": f"{latest_aum.confidence_score:.1%}" if latest_aum else "0%",
                "Scrapes Bem-sucedidos": successful_scrapes,
                "Scrapes Falharam": failed_scrapes,
                "URLs Scrapadas": "; ".join(scraped_urls[:3]) if scraped_urls else "Nenhuma",
                "Data Extração": latest_aum.extracted_at.strftime('%d/%m/%Y %H:%M') if latest_aum else "N/A",
                "Empresa Criada": company.created_at.strftime('%d/%m/%Y %H:%M')
            }
            
            data.append(row)
        
        df = pd.DataFrame(data)
        
        # Ordenar por AUM (maiores primeiro)
        df['AUM_Sort'] = df['AUM Normalizado'].fillna(0)
        df = df.sort_values('AUM_Sort', ascending=False)
        df = df.drop('AUM_Sort', axis=1)
        
        return df
    
    async def _create_logs_dataframe(self, db: AsyncSession) -> pd.DataFrame:
        """Cria DataFrame com logs de scraping"""
        
        result = await db.execute(
            select(ScrapeLog)
            .join(Company)
            .order_by(ScrapeLog.scraped_at.desc())
        )
        
        logs = result.scalars().all()
        
        data = []
        for log in logs:
            data.append({
                "Empresa": log.company.name,
                "URL": log.url,
                "Tipo Conteúdo": log.content_type,
                "Status": log.status,
                "Erro": log.error_message or "N/A",
                "Tamanho Conteúdo": len(log.scraped_content) if log.scraped_content else 0,
                "Data Scraping": log.scraped_at.strftime('%d/%m/%Y %H:%M:%S')
            })
        
        return pd.DataFrame(data)
    
    async def _create_summary_dataframe(self, companies_data: List[Dict]) -> pd.DataFrame:
        """Cria DataFrame com resumo estatístico"""
        
        # Estatísticas gerais
        total_companies = len(companies_data)
        companies_with_aum = len([item for item in companies_data 
                                 if item["latest_aum"] and item["latest_aum"].aum_normalized])
        
        aum_values = [item["latest_aum"].aum_normalized 
                     for item in companies_data 
                     if item["latest_aum"] and item["latest_aum"].aum_normalized]
        
        data = [
            {"Métrica": "Total de Empresas", "Valor": total_companies},
            {"Métrica": "Empresas com AUM Encontrado", "Valor": companies_with_aum},
            {"Métrica": "Taxa de Sucesso", "Valor": f"{companies_with_aum/total_companies:.1%}" if total_companies > 0 else "0%"},
            {"Métrica": "AUM Total (R$)", "Valor": format_currency(sum(aum_values)) if aum_values else "N/A"},
            {"Métrica": "AUM Médio (R$)", "Valor": format_currency(sum(aum_values)/len(aum_values)) if aum_values else "N/A"},
            {"Métrica": "Maior AUM (R$)", "Valor": format_currency(max(aum_values)) if aum_values else "N/A"},
            {"Métrica": "Menor AUM (R$)", "Valor": format_currency(min(aum_values)) if aum_values else "N/A"},
            {"Métrica": "Data Geração", "Valor": datetime.now().strftime('%d/%m/%Y %H:%M:%S')}
        ]
        
        return pd.DataFrame(data)
    
    async def _format_excel_sheets(self, writer):
        """Formata as planilhas Excel"""
        
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            workbook = writer.book
            
            # Formatar planilha principal
            if 'Resultados AUM' in workbook.sheetnames:
                ws = workbook['Resultados AUM']
                
                # Header styling
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Formatar planilha de resumo
            if 'Resumo' in workbook.sheetnames:
                ws = workbook['Resumo']
                
                # Header styling
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                
                # Auto-adjust column widths
                ws.column_dimensions['A'].width = 25
                ws.column_dimensions['B'].width = 20
                
        except Exception as e:
            logger.warning(f"Erro ao formatar Excel: {str(e)}")
    
    def get_filename(self) -> str:
        """Retorna nome do arquivo Excel"""
        return self.output_filename